import pandas as pd
import cplex
from readershappyinstances import GraphInstance
import readershappyinstances
import networkx as nx
import time
import os
from docplex.cp.model import CpoModel
from docplex.mp.model import Model
import math
import logging
import openpyxl 
import sys
from contextlib import contextmanager

# Função para configurar o logger
def setup_logger(name, log_file, level=logging.INFO):
    """Configura o logger para gravar mensagens em um arquivo específico."""
    logger = logging.getLogger(name)
    handler = logging.FileHandler(log_file)
    handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.setLevel(level)
    logger.addHandler(handler)
    return logger

# Context Manager para redirecionar stdout e stderr para um arquivo de log
@contextmanager
def redirect_output(log_file):
    """Context manager para redirecionar stdout e stderr para um arquivo de log."""
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    with open(log_file, 'w') as log:
        sys.stdout = log
        sys.stderr = log
        try:
            yield
        finally:
            sys.stdout = original_stdout
            sys.stderr = original_stderr

def maxhs_cp(graph, k, log_file):
    cpomdl = CpoModel()
    n = len(graph)
    x = cpomdl.binary_var_list(n, name="x")
    objective = cpomdl.sum(
        cpomdl.min([x[j] for j in (graph[i] + [i])])  # mínimo entre os vizinhos de i + i
        for i in range(n)
    )
    cpomdl.maximize(objective)
    cpomdl.add_constraint(cpomdl.sum(x) == k)

    cpoptimizer_path = '/home/rafael/CPLEX_Studio2211/cpoptimizer/bin/x86-64_linux/cpoptimizer'

    start_time = time.time()  # Tempo de início da execução

    # Criando o logger para capturar as saídas no log
    logger = setup_logger('CPOptimizer', log_file)

    # Abrindo o arquivo de log corretamente
    with open(log_file, 'w') as log_output:
        logger.info("Iniciando a otimização com CP Optimizer.")

        # Passando o objeto de arquivo log_output para o parâmetro log_output
        solution = cpomdl.solve(
            TimeLimit=60, 
            Presolve='Off', 
            Workers=1, 
            execfile=cpoptimizer_path, 
            log_output=log_output,
            OptimalityTolerance=0.01 
        )
    
    exec_time = time.time() - start_time  # Tempo de execução

    if solution:
        selected_vertices = [i for i in range(n) if solution[x[i]] == 1]
        happy_vertices = []
        for i in range(n):
            closed_neighbors = graph[i] + [i]  # Vizinhos de i, incluindo o próprio vértice
            if all(solution[x[j]] == 1 for j in closed_neighbors):
                happy_vertices.append(i)
        objective_value = solution.get_objective_value()  # Valor da função objetivo
        gap_cp = solution.get_objective_gap() 
        return selected_vertices, happy_vertices, exec_time, objective_value, gap_cp
    else:
        logger.error("Solução não encontrada!")
        return None, None, exec_time, None

def solve_original_model(lista_viz, k, log_file):
    mdl = Model(log_output=True)
    n = len(lista_viz)
    y = mdl.binary_var_list(n, name="y")
    h = mdl.continuous_var_list(n, name="h")

    mdl.maximize(mdl.sum(h[i] for i in range(n)))
    mdl.add_constraint(mdl.sum(y) == k)
    for i in range(n):
        mdl.add_constraint(h[i] <= y[i])
    for i in range(n):
        for j in lista_viz[i]:  # j ∈ Γ(i)
            mdl.add_constraint(h[i] <= y[j])
    for i in range(n):
        mdl.add_constraint(h[i] >= mdl.sum(y[j] for j in lista_viz[i]) - len(lista_viz[i]) + y[i])

    mdl.parameters.mip.strategy.heuristicfreq.set(-1)
    mdl.parameters.mip.cuts.mircut.set(-1)
    mdl.parameters.mip.cuts.implied.set(-1)
    mdl.parameters.mip.cuts.gomory.set(-1)
    mdl.parameters.mip.cuts.flowcovers.set(-1)
    mdl.parameters.mip.cuts.pathcut.set(-1)
    mdl.parameters.mip.cuts.liftproj.set(-1)
    mdl.parameters.mip.cuts.zerohalfcut.set(-1)
    mdl.parameters.mip.cuts.cliques.set(-1)
    mdl.parameters.mip.cuts.covers.set(-1)
    mdl.parameters.threads.set(1)
    mdl.parameters.clocktype.set(1)
    mdl.parameters.timelimit.set(60)
    mdl.parameters.preprocessing.presolve.set(0)
    mdl.parameters.preprocessing.boundstrength.set(0)
    mdl.parameters.preprocessing.coeffreduce.set(0)
    mdl.parameters.preprocessing.relax.set(0)
    mdl.parameters.preprocessing.aggregator.set(0)
    mdl.parameters.preprocessing.reduce.set(0)
    mdl.parameters.preprocessing.reformulations.set(0)

    # Criando o logger para capturar as saídas no log
    logger = setup_logger('OriginalModel', log_file)
    with redirect_output(log_file):
        logger.info("Iniciando a otimização com o modelo original.")

        start_time = time.time()  # Tempo de início da execução
        solution = mdl.solve()
        exec_time = time.time() - start_time  # Tempo de execução

    if solution:
        selected_vertices = [i for i in range(n) if solution.get_value(y[i]) == 1]
        happy_vertices = [i for i in range(n) if solution.get_value(h[i]) == 1]
        objective_value = solution.objective_value  # Valor da função objetivo
        return selected_vertices, happy_vertices, exec_time, objective_value
    else:
        logger.error("Solução não encontrada!")
        return None, None, exec_time, None


def model_benders(lista_viz, k):
    mdl = Model(log_output=True)
    n = len(lista_viz)
    y = mdl.binary_var_list(n, name="y")
    h = mdl.continuous_var_list(n, name="h")

    mdl.maximize(mdl.sum(h[i] for i in range(n)))
    mdl.add_constraint(mdl.sum(y) == k)
    for i in range(n):
        mdl.add_constraint(h[i] <= y[i])
    for i in range(n):
        for j in lista_viz[i]:  # j ∈ Γ(i)
            mdl.add_constraint(h[i] <= y[j])
    for i in range(n):
        mdl.add_constraint(h[i] >= mdl.sum(y[j] for j in lista_viz[i]) - len(lista_viz[i]) + y[i])

    mdl.parameters.benders.strategy = 3
    mdl.parameters.mip.strategy.heuristicfreq.set(-1)
    mdl.parameters.mip.cuts.mircut.set(-1)
    mdl.parameters.mip.cuts.implied.set(-1)
    mdl.parameters.mip.cuts.gomory.set(-1)
    mdl.parameters.mip.cuts.flowcovers.set(-1)
    mdl.parameters.mip.cuts.pathcut.set(-1)
    mdl.parameters.mip.cuts.liftproj.set(-1)
    mdl.parameters.mip.cuts.zerohalfcut.set(-1)
    mdl.parameters.mip.cuts.cliques.set(-1)
    mdl.parameters.mip.cuts.covers.set(-1)
    mdl.parameters.threads.set(1)
    mdl.parameters.clocktype.set(1)
    mdl.parameters.timelimit.set(60)
    mdl.parameters.preprocessing.presolve.set(0)
    mdl.parameters.preprocessing.boundstrength.set(0)
    mdl.parameters.preprocessing.coeffreduce.set(0)
    mdl.parameters.preprocessing.relax.set(0)
    mdl.parameters.preprocessing.aggregator.set(0)
    mdl.parameters.preprocessing.reduce.set(0)
    mdl.parameters.preprocessing.reformulations.set(0)

    # # Criando o logger para capturar as saídas no log
    # logger = setup_logger('OriginalModel', log_file)
    # with redirect_output(log_file):
    #     logger.info("Iniciando a otimização com o modelo original.")

    start_time = time.time()  # Tempo de início da execução
    solution = mdl.solve()
    exec_time = time.time() - start_time  # Tempo de execução

    if solution:
        selected_vertices = [i for i in range(n) if solution.get_value(y[i]) == 1]
        happy_vertices = [i for i in range(n) if solution.get_value(h[i]) == 1]
        objective_value = solution.objective_value  # Valor da função objetivo
        return selected_vertices, happy_vertices, exec_time, objective_value
    else:
        logger.error("Solução não encontrada!")
        return None, None, exec_time, None



# Função para salvar resultados no Excel na primeira linha disponível
def save_results_to_excel(results, log_folder, filename="results_comparison_cp_benders_results_gap.xlsx"):
    excel_file_path = os.path.join(log_folder, filename)

    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        row = sheet.max_row + 1
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["instance_name", "k", "selected_vertices_cp", "happy_vertices_cp", 
                      "exec_time_cp", "objective_value_cp", "gap_cp",
                      "selected_vertices_orig", "happy_vertices_orig", 
                      "exec_time_orig", "objective_value_orig", 
                      "selected_vertices_ben_aut", "happy_vertices_ben_aut", 
                      "exec_time_ben_aut", "objective_value_ben_aut"])
        row = 2

    for result in results:
        # Converte as listas para strings
        selected_vertices_cp_str = result["selected_vertices_cp"]
        happy_vertices_cp_str = result["happy_vertices_cp"]
        selected_vertices_orig_str = result["selected_vertices_orig"]
        happy_vertices_orig_str = result["happy_vertices_orig"]
        selected_vertices_ben_aut_str= result["selected_vertices_ben_aut"]
        happy_vertices_ben_aut_str= result["happy_vertices_ben_aut"]

        # happy_vertices_cp_str = ','.join(map(str, result["happy_vertices_cp"]))
        # selected_vertices_orig_str = ','.join(map(str, result["selected_vertices_orig"]))
        # happy_vertices_orig_str = ','.join(map(str, result["happy_vertices_orig"]))
        # selected_vertices_ben_aut_str= ','.join(map(str, result["selected_vertices_ben_aut"]))
        # happy_vertices_ben_aut_str= ','.join(map(str, result["happy_vertices_ben_aut"]))


        sheet.append([result["instance_name"], result["k"], selected_vertices_cp_str,
                      happy_vertices_cp_str, result["exec_time_cp"], result["objective_value_cp"], result["gap_cp"],
                      selected_vertices_orig_str, happy_vertices_orig_str,
                      result["exec_time_orig"], result["objective_value_orig"],
                      selected_vertices_ben_aut_str, happy_vertices_ben_aut_str,
                      result["exec_time_ben_aut"], result["objective_value_ben_aut"]])
        
        

    workbook.save(excel_file_path)
    print(f"Resultados salvos em {excel_file_path}")


# Função para criar a pasta de logs
def create_log_folder():
    log_folder = '/home/rafael/Downloads/CP'
    if not os.path.exists(log_folder):
        os.makedirs(log_folder)  # Cria a pasta se não existir
    return log_folder

# Função para calcular os valores de k a serem testados
def get_k_values(k, num_vertices):
    k_values = [
        # math.ceil(0.20 * num_vertices), 
        math.ceil(0.50 * num_vertices), 
        # math.ceil(0.90 * num_vertices),
        num_vertices
    ]
    
    # Filtrar os valores para manter apenas os maiores que o valor de k
    k_values = [value for value in k_values if value > k]

    # Adiciona k na lista, caso seja maior que 0 e não tenha sido descartado
    if k > 0 and k not in k_values:
        k_values.insert(0, k)

    return sorted(set(k_values))  

# Função para obter arquivos em uma pasta com uma extensão específica
def get_files_in_folder(folder_path, extension=".txt"):
    files = [f for f in os.listdir(folder_path) if f.endswith(extension)]
    full_paths = [os.path.join(folder_path, f) for f in files]
    return full_paths

# Função principal para processar os arquivos
def process_files_in_folder(folder_path):
    files = get_files_in_folder(folder_path)
    results = []

    # Criar a pasta de logs, caso não exista
    log_folder = create_log_folder()

    for file in files:
        print(f"\nProcessando arquivo: {file}")
        instancia = GraphInstance(file)
        adj_list, precolored, vertices, arestas, k = instancia.get_data()
        print(adj_list)
        print(k)

        # Nome da instância (arquivo sem extensão)
        instance_name = os.path.basename(file)

        # Calcular os valores de k a serem testados
        k_values = get_k_values(k, len(vertices))

        # Resolver o problema para os valores de k calculados
        for k_val in k_values:
            print(f"\nValor de k: {k_val}")
            
            # Definir o nome do arquivo de log
            log_file_path = os.path.join(log_folder, f'log_{instance_name}_k{k_val}.txt')
            
            # Algoritmo CP Optimizer
            selected_vertices_cp, happy_vertices_cp, exec_time_cp, objective_value_cp, gap = maxhs_cp(adj_list, k_val, log_file_path)
            print("Algoritmo CP Optimizer:")
            print("Vértices selecionados:", selected_vertices_cp)
            print("Vértices felizes:", happy_vertices_cp)
            print(f"Tempo de execução: {exec_time_cp:.4f} segundos")
            print(f"Valor da função objetivo: {objective_value_cp}")
            print(f"Valor do GAP do CP: {gap}")

            # Algoritmo Original
            selected_vertices_orig, happy_vertices_orig, exec_time_orig, objective_value_orig = solve_original_model(adj_list, k_val, log_file_path)
            print("Algoritmo Original:")
            print("Vértices selecionados:", selected_vertices_orig)
            print("Vértices felizes:", happy_vertices_orig)
            print(f"Tempo de execução: {exec_time_orig:.4f} segundos")
            print(f"Valor da função objetivo: {objective_value_orig}")

             # Benders
            selected_vertices_ben_aut, happy_vertices_ben_aut, exec_time_ben_aut, objective_value_ben_aut = model_benders(adj_list, k_val)
            print("Algoritmo Automatico de Benders:")
            print("Vértices selecionados:", selected_vertices_ben_aut)
            print("Vértices felizes:", happy_vertices_ben_aut)
            print(f"Tempo de execução: {exec_time_ben_aut:.4f} segundos")
            print(f"Valor da função objetivo: {objective_value_ben_aut}")

            # Armazenando os resultados como listas (strings)
            results.append({
                "instance_name": instance_name,
                "k": k_val,
                "selected_vertices_cp": str(selected_vertices_cp),
                "happy_vertices_cp": str(happy_vertices_cp),
                "exec_time_cp": exec_time_cp,
                "objective_value_cp": objective_value_cp,
                "gap_cp":gap,
                "selected_vertices_orig": str(selected_vertices_orig),
                "happy_vertices_orig": str(happy_vertices_orig),
                "exec_time_orig": exec_time_orig,
                "objective_value_orig": objective_value_orig,
                "selected_vertices_ben_aut": str(selected_vertices_ben_aut),
                "happy_vertices_ben_aut": str(happy_vertices_ben_aut),
                "exec_time_ben_aut": exec_time_ben_aut,
                "objective_value_ben_aut": objective_value_ben_aut
            })

    # Salvar os resultados em um arquivo Excel dentro da pasta CP
    save_results_to_excel(results, log_folder)

# Caminho da pasta com os arquivos
folder_path = '/home/rafael/Documents/HappySet/MIHS/inputs/happygen/output/testes/testes_cp_benders'
# folder_path ='/home/rafael/Documents/HappySet/MIHS/inputs/happygen/output/testes/7-10'

# Chama a função para processar os arquivos da pasta
process_files_in_folder(folder_path)
